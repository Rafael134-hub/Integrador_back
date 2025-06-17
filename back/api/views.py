from django.shortcuts import render
from .models import *
from .serializer import *
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework import generics
from django.contrib.auth.models import User
from rest_framework.parsers import MultiPartParser
from rest_framework.views import APIView
from openpyxl import load_workbook, Workbook
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend
from django.shortcuts import get_object_or_404
from .filters import *
from django.http import HttpResponse
from datetime import datetime

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def listar_sensores(request):
    if request.method == 'GET':
        queryset = Sensor.objects.all()
        serializer = Sensor_serializer(queryset, many=True)
        return Response(serializer.data)
    elif request.method == 'POST':
        serializer = Sensor_serializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)

class SensoresView(ListCreateAPIView):
    queryset = Sensor.objects.all()
    serializer_class = Sensor_serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = Sensores_filter

class SensoresDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Sensor.objects.all()
    serializer_class = Sensor_serializer
    permission_classes = [IsAuthenticated]
    


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def listar_ambientes(request):
    if request.method == 'GET':
        queryset = Ambiente.objects.all()
        serializer = Ambiente_serializer(queryset, many=True)
        return Response(serializer.data)
    elif request.method == 'POST':
        serializer = Ambiente_serializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)

class AmbientesView(ListCreateAPIView):
    queryset = Ambiente.objects.all()
    serializer_class = Ambiente_serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = Ambientes_filter

class AmbientesDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Ambiente.objects.all()
    serializer_class = Ambiente_serializer
    permission_classes = [IsAuthenticated]



@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def listar_historicos(request):
    if request.method == 'GET':
        queryset = Historico.objects.select_related("id_sensor", "id_ambiente").all()
        serializer = Historico_serializer(queryset, many=True)
        return Response(serializer.data)
    elif request.method == 'POST':
        serializer = Historico_serializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)

class HistoricosView(ListCreateAPIView):
    queryset = Historico.objects.select_related("id_sensor", "id_ambiente").all()
    serializer_class = Historico_serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = Historicos_filter
    

class HistoricosDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Historico.objects.select_related("id_sensor", "id_ambiente").all()
    serializer_class = Historico_serializer
    permission_classes = [IsAuthenticated]

class SignUpView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]

class UploadXLSXViewAmbiente(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  
        
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)): 
            sig = str(row[0]).strip() if row[0] is not None else None
            descricao = str(row[1]).strip() if row[1] is not None else None
            ni = str(row[2]).strip() if row[2] is not None else None
            responsavel = str(row[3]).strip() if row[3] is not None else None
            
            if not ni:
                    print(f"[Linha {i+2}] Erro: ni vazio. Dados: {row}")
                
            Ambiente.objects.create(
                sig=row[0],
                descricao=row[1],
                ni=row[2],
                responsavel=row[3],
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})
    

class UploadXLSXViewSensores(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  # primeira aba
            
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # pula o cabeçalho
            sensor = str(row[0]).strip() if row[0] is not None else None
            mac_adress = str(row[1]).strip() if row[1] is not None else None
            unidade_med = str(row[2]).strip() if row[2] is not None else None
            latitude = str(row[3]).strip() if row[3] is not None else None
            longitude = str(row[4]).strip() if row[4] is not None else None
            status = str(row[5]).strip() if row[5] is not None else None
                
            if not sensor:
                    print(f"[Linha {i+2}] Erro: sensor vazio. Dados: {row}")
                    
            Sensor.objects.create(
                sensor = row[0],
                mac_adress = row[1],
                unidade_med = row[2],
                latitude = row[3],
                longitude = row[4],
                status = bool(row[5]),
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})
    
    
class UploadXLSXViewHistoricos(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  
            
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  
            id_ambiente = str(row[0]).strip() if row[0] is not None else None
            id_sensor = str(row[1]).strip() if row[1] is not None else None
            valor = str(row[2]).strip() if row[2] is not None else None
            timestamp = str(row[3]).strip() if row[3] is not None else None
            
                
            if not valor:
                    print(f"[Linha {i+2}] Erro: sensor vazio. Dados: {row}")
                    
            try:
                id_ambiente = get_object_or_404(Ambiente, id=int(row[0]))
                id_sensor = get_object_or_404(Sensor, id=int(row[1]))
                valor = float(row[2])
                timestamp = row[3]

                Historico.objects.create(
                    id_ambiente=id_ambiente,
                    id_sensor=id_sensor,
                    timestamp=timestamp,
                    valor=valor
                )

            except Exception as e:
                print(f"[Linha {i+2}] Erro ao processar linha: {e}. Dados: {row}")

        return Response({'mensagem': 'Dados importados com sucesso!'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_sensores_excel(request):
    filterset = Sensores_filter(request.GET, queryset=Sensor.objects.all())
    sensores = filterset.qs

    wb = Workbook()
    ws = wb.active
    ws.title = "Sensores"

    ws.append([
        "ID", "Sensor", "MAC Address", "Unidade de Medida",
        "Latitude", "Longitude", "Status"
    ])

    for s in sensores:
        ws.append([
            s.id,
            s.sensor,
            s.mac_adress,
            s.unidade_med,
            s.latitude,
            s.longitude,
            "Ativo" if s.status else "Inativo"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=sensores.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_ambientes_excel(request):
    filterset = Ambientes_filter(request.GET, queryset = Ambiente.objects.all())
    ambientes = filterset.qs

    wb = Workbook()
    ws = wb.active
    ws.title = "Ambientes"

    ws.append([
        "ID", "SIG", "Descrição", "NI",
        "Responsável"
    ])

    for a in ambientes:
        ws.append([
            a.id,
            a.sig,
            a.descricao,
            a.ni,
            a.responsavel
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ambientes.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_historicos_excel(request):
    filterset = Historicos_filter(request.GET, queryset=Historico.objects.all())
    historicos = filterset.qs

    wb = Workbook()
    ws = wb.active
    ws.title = "Históricos"

    ws.append([
        "ID", "Sensor", "Ambiente", "Valor",
        "Timestamp"
    ])

    for h in historicos:
        try:
            timestamp_str = str(h.timestamp)
            
            if len(timestamp_str) < 14:
                timestamp_str = timestamp_str.zfill(14)
            
            timestamp_formatado = timestamp_str[0:2] + "/" + timestamp_str[2:4] + "/" + timestamp_str[4:8] + " " + timestamp_str[8:10] + ":" + timestamp_str[10:12] + ":" + timestamp_str[12:14] 
            
        except Exception as e:
            timestamp_formatado = "Erro no timestamp"

        ws.append([
            h.id,
            h.id_sensor.sensor if h.id_sensor else "",
            h.id_ambiente.descricao if h.id_ambiente else "",
            h.valor,
            timestamp_formatado
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historicos.xlsx'
    wb.save(response)
    return response